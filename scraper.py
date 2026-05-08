"""
Safespill Hangar News Scraper
------------------------------
Runs weekly, searches for new hangar/fire-protection project news across
North America and EMEA using SerpAPI (Google Search), SAM.gov, CanadaBuys,
AusTender, and TED Europa. Emails a formatted Excel report with two tabs
(North America and EMEA) to the Safespill team every Monday at 8am PST.

Setup: see README.md
"""

import os
import io
import re
import time
import datetime
import logging
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from html.parser import HTMLParser

# --- Logging -----------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger(__name__)

# --- Configuration -----------------------------------------------------------
SERPAPI_KEY   = os.environ["SERPAPI_KEY"]
SMTP_HOST     = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT     = int(os.environ.get("SMTP_PORT", 587))
SMTP_USER     = os.environ["SMTP_USER"]
SMTP_PASSWORD = os.environ["SMTP_PASSWORD"]
RECIPIENT     = os.environ.get("REPORT_RECIPIENT", "trevorw@safespill.com")

# --- Search configuration ----------------------------------------------------
#
# Query strategy: all queries are run globally and articles get sorted into
# the correct tab by detect_location() afterward. Queries are split into two
# lists for readability — there is no "NA list / EMEA list" anymore since
# the previous overlap was creating duplicate API calls for the same results.
#
# GENERAL_QUERIES: broad aviation / fire-protection terms that could surface
#   articles from any region. These do the bulk of the discovery work.
#
# REGIONAL_QUERIES: include country or city names to specifically surface
#   projects in EMEA / APAC markets that the general queries miss.

GENERAL_QUERIES = [
    # New construction & development
    "aircraft hangar construction",
    "new aircraft hangar",
    "hangar development project",
    "airport hangar expansion",
    "hangar construction contract awarded",
    "hangar groundbreaking",
    # FBO / private / corporate aviation
    "FBO hangar development",
    "FBO expansion",
    "private aviation hangar",
    "corporate hangar construction",
    "business jet hangar",
    # Retrofit & renovation
    "hangar retrofit project",
    "hangar renovation contract",
    # MRO & maintenance facilities
    "MRO facility construction",
    "aircraft maintenance hangar",
    "MRO hangar project",
    "wide-body maintenance hangar",
    "narrow-body MRO",
    "engine shop hangar",
    "paint hangar construction",
    "completions center aviation",
    # Military
    "military hangar construction",
    "Air Force hangar project",
    "Navy hangar construction",
    "MILCON hangar contract",
    # Fire protection — direct
    "hangar fire protection system",
    "aircraft hangar fire suppression",
    "hangar foam suppression",
    # Fire protection — PFAS / AFFF transition (drives retrofit demand)
    "AFFF hangar replacement",
    "fluorine free foam hangar",
    "NFPA 409 hangar",
]

# Site filters removed — free SerpAPI tier works best with clean short queries

NA_COUNTRIES = {"United States", "Canada", "Mexico"}

SAM_KEYWORDS = [
    "hangar", "aircraft", "airfield", "aviation", "fire suppression",
    "fire protection", "foam", "suppression system",
]

LOOKBACK_DAYS = 7

# SAM.gov uses a longer lookback since contracts are posted infrequently
SAM_LOOKBACK_DAYS = 30

# --- Regional (location-specific) queries -----------------------------------
# These include country or city names. They mostly target EMEA markets that
# the general queries tend to under-surface.

REGIONAL_QUERIES = [
    # Middle East
    "hangar construction Dubai",
    "hangar project UAE",
    "aircraft hangar Saudi Arabia",
    "MRO facility Middle East",
    "hangar Qatar",
    # Africa
    "aircraft hangar Nigeria",
    "MRO facility Africa",
    "MRO hangar Kenya",
    "aviation facility South Africa",
    # Europe
    "hangar construction UK",
    "aircraft maintenance hangar Germany",
    "hangar construction Poland",
    "aircraft hangar Turkey",
    # Defense (NATO / non-US military)
    "NATO hangar construction",
    "defence hangar project",
]

# EMEA news domains extracted from known relevant sources
EMEA_NEWS_DOMAINS = [
    "zawya.com", "adsadvance.co.uk", "aviationbusinessme.com",
    "asianaviation.com", "arabianbusiness.com", "mepmiddleeast.com",
    "aviationweek.com", "aviationpros.com", "ainonline.com",
    "aerotime.aero", "aviationsourcenews.com", "avitrader.com",
    "engineeringnews.co.za", "ecofinagency.com", "businessairportinternational.com",
    "aircraftinteriorsinternational.com", "traveldailynews.com", "agbi.com",
    "breakingdefense.com", "defence-industry.eu", "gulfbusiness.com",
    "thisdaylive.com", "vanguardngr.com", "newtelegraphng.com",
    "punchng.com", "thesun.ng", "legit.ng", "arise.tv", "apanews.net",
    "addisinsight.net", "united24media.com", "haaretz.com",
    "ted.europa.eu", "travelandtourworld.com", "aviationbusinessnews.com",
]



HEADERS_UA = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

# --- Helpers -----------------------------------------------------------------

def date_range():
    end   = datetime.date.today()
    start = end - datetime.timedelta(days=LOOKBACK_DAYS)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def week_label():
    today  = datetime.date.today()
    monday = today - datetime.timedelta(days=today.weekday())
    return monday.strftime("%Y-%m-%d")


# --- Article page fetching ---------------------------------------------------

# Common <meta> date property names used by news sites
META_DATE_PROPS = [
    "article:published_time",
    "article:modified_time",
    "og:published_time",
    "datePublished",
    "date",
    "pubdate",
    "publish_date",
    "DC.date",
    "DC.Date",
]

# Common JSON-LD / schema.org date fields
JSONLD_DATE_KEYS = ["datePublished", "dateCreated", "dateModified"]


class MetaParser(HTMLParser):
    """Extract <meta> tags, <title>, and <script type=application/ld+json> from HTML."""

    def __init__(self):
        super().__init__()
        self.meta    = {}   # property/name -> content
        self.scripts = []   # raw text of ld+json blocks
        self.title   = ""
        self._in_jsonld = False
        self._in_title  = False
        self._buf       = []
        self._title_buf = []

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "meta":
            key   = attrs.get("property") or attrs.get("name") or ""
            value = attrs.get("content", "")
            if key and value:
                self.meta[key.lower()] = value
        if tag == "script" and attrs.get("type") == "application/ld+json":
            self._in_jsonld = True
            self._buf = []
        if tag == "title":
            self._in_title  = True
            self._title_buf = []

    def handle_endtag(self, tag):
        if tag == "script" and self._in_jsonld:
            self.scripts.append("".join(self._buf))
            self._in_jsonld = False
            self._buf = []
        if tag == "title" and self._in_title:
            if not self.title:  # only first <title>
                self.title = "".join(self._title_buf).strip()
            self._in_title  = False
            self._title_buf = []

    def handle_data(self, data):
        if self._in_jsonld:
            self._buf.append(data)
        if self._in_title:
            self._title_buf.append(data)


def fetch_article_meta(url: str) -> dict:
    """
    Fetch an article page and extract:
      - publish_date  (YYYY-MM-DD string or "")
      - location      (str or "")  -- detected country/region
    Returns a dict with those keys.
    """
    result = {"publish_date": "", "location": ""}
    try:
        r = requests.get(url, headers=HEADERS_UA, timeout=15, allow_redirects=True)
        if r.status_code != 200:
            return result
        html = r.text[:150_000]   # only parse first ~150 KB to stay fast
    except Exception as exc:
        log.debug("Could not fetch %s: %s", url, exc)
        return result

    parser = MetaParser()
    try:
        # FIX: HTMLParser uses .feed(), not .parse(). The previous code silently
        # failed and meta/JSON-LD data was never extracted.
        parser.feed(html)
    except Exception:
        pass

    # --- Extract date --------------------------------------------------------
    date_str = ""

    # 1. Check <meta> tags
    for prop in META_DATE_PROPS:
        val = parser.meta.get(prop.lower(), "")
        if val:
            date_str = val
            break

    # 2. Check JSON-LD blocks
    if not date_str:
        import json
        for script in parser.scripts:
            try:
                data = json.loads(script)
                # data might be a list
                items = data if isinstance(data, list) else [data]
                for item in items:
                    for key in JSONLD_DATE_KEYS:
                        if item.get(key):
                            date_str = item[key]
                            break
                    if date_str:
                        break
            except Exception:
                continue

    # 3. Regex fallback — look for ISO date in the HTML
    if not date_str:
        m = re.search(r'(\d{4}-\d{2}-\d{2})', html[:50_000])
        if m:
            date_str = m.group(1)

    if date_str:
        result["publish_date"] = parse_iso_date(date_str)

    # --- Extract location ----------------------------------------------------
    # Use only the title and the article opening (~3000 chars of visible text)
    # rather than the entire page. This avoids footer/nav/sidebar pollution
    # where unrelated locations from a site's other content cause misclassification.
    title_text = parser.title or parser.meta.get("og:title", "") or ""
    desc_text  = parser.meta.get("og:description", "") or parser.meta.get("description", "") or ""

    visible = re.sub(r'<[^>]+>', ' ', html)
    # Collapse whitespace
    visible = re.sub(r'\s+', ' ', visible)
    # Limit body text to first 3000 chars (article opening)
    body_snippet = (desc_text + " " + visible[:3000]).strip()

    result["location"] = detect_location(body_snippet, title=title_text)

    return result


def parse_iso_date(raw: str) -> str:
    """Normalise various date string formats to YYYY-MM-DD."""
    raw = raw.strip()
    # ISO 8601 with time component e.g. 2025-12-03T14:30:00Z
    m = re.match(r'(\d{4}-\d{2}-\d{2})', raw)
    if m:
        return m.group(1)
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%m/%d/%Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


# --- SerpAPI quota tracking --------------------------------------------------

SERPAPI_ACCOUNT_URL = "https://serpapi.com/account.json"


def check_serpapi_quota() -> dict:
    """
    Query SerpAPI's free /account.json endpoint to check quota usage.
    This call is FREE — it does not count against your monthly search budget.
    Returns a dict with quota info, or an empty dict if the call fails.
    """
    try:
        r = requests.get(
            SERPAPI_ACCOUNT_URL,
            params={"api_key": SERPAPI_KEY},
            timeout=10,
        )
        r.raise_for_status()
        return r.json()
    except Exception as exc:
        log.warning("Could not check SerpAPI quota: %s", exc)
        return {}


def log_quota_status(label: str, quota: dict):
    """Pretty-print SerpAPI quota info to the log."""
    if not quota:
        log.info("[%s] SerpAPI quota: <unable to fetch>", label)
        return
    plan       = quota.get("plan_name", "unknown")
    left       = quota.get("total_searches_left", "?")
    monthly    = quota.get("searches_per_month", "?")
    used       = quota.get("this_month_usage", "?")
    hourly     = quota.get("last_hour_searches", "?")
    hourly_cap = quota.get("account_rate_limit_per_hour", "?")
    log.info(
        "[%s] SerpAPI quota — plan=%s | %s/%s used this month | "
        "%s left | %s/%s used last hour",
        label, plan, used, monthly, left, hourly, hourly_cap,
    )


# --- SerpAPI -----------------------------------------------------------------

# Counts SerpAPI rate-limit / quota errors during a run. Reset at start of
# main(). Used to add a warning notice to the team email if errors occurred.
_SERPAPI_ERROR_COUNT = 0


def serpapi_news_search(query: str) -> list[dict]:
    global _SERPAPI_ERROR_COUNT
    url = "https://serpapi.com/search"
    params = {
        "engine":  "google",
        "q":       query,
        "gl":      "us",
        "hl":      "en",
        # SerpAPI bills per search, NOT per result. Bumping num from 10 → 100
        # gets up to 10x more results at zero additional API cost.
        "num":     100,
        "api_key": SERPAPI_KEY,
    }
    try:
        r = requests.get(url, params=params, timeout=20)

        # Detect rate-limit / quota errors explicitly so they show up clearly
        # in logs instead of looking like an empty-result query.

        # Hourly throughput cap exceeded (HTTP 429)
        if r.status_code == 429:
            _SERPAPI_ERROR_COUNT += 1
            log.error(
                "SerpAPI RATE LIMITED for '%s' (HTTP 429). "
                "Hourly throughput cap reached (free tier = 50/hour). "
                "Either you re-triggered a run too soon after the previous one, "
                "or you're at the monthly cap. Check https://serpapi.com/dashboard.",
                query,
            )
            return []

        r.raise_for_status()
        data = r.json()

        # SerpAPI sometimes returns HTTP 200 with an 'error' field for
        # quota / auth / parameter problems. Surface these explicitly.
        if "error" in data:
            err_msg   = data.get("error", "")
            err_lower = err_msg.lower()
            if any(term in err_lower for term in ("run out", "exceeded", "limit", "quota")):
                _SERPAPI_ERROR_COUNT += 1
                log.error(
                    "SerpAPI QUOTA ERROR for '%s': %s. "
                    "Likely exhausted your monthly free searches (250/month). "
                    "Reduce query count or upgrade plan.",
                    query, err_msg,
                )
            else:
                log.warning("SerpAPI returned error for '%s': %s", query, err_msg)
            return []

        results = data.get("organic_results", [])
        log.info("SerpAPI '%s' -> %d results", query, len(results))
        return results
    except Exception as exc:
        log.warning("SerpAPI error for '%s': %s", query, exc)
        return []


def parse_serpapi_result(item: dict) -> dict | None:
    title   = item.get("title", "").strip()
    link    = item.get("link", "").strip()
    snippet = item.get("snippet", "").strip()

    if not title or not link:
        return None

    # Fallback date/location from snippet (will be overwritten by page fetch)
    date_raw      = item.get("date", "")
    fallback_date = parse_google_date(date_raw)
    # Pass title separately so detect_location can weight headline mentions higher
    location      = detect_location(snippet, title=title)

    return {
        "Project Title":  title,
        "Source URL":     link,
        "Summary":        snippet,
        "Date Published": fallback_date,
        "Location":       location,
        "Source":         "news",
    }


def parse_google_date(raw: str) -> str:
    raw   = raw.strip().lower()
    today = datetime.date.today()
    if "hour" in raw or "minute" in raw or "second" in raw:
        return today.strftime("%Y-%m-%d")
    if "day" in raw:
        try:
            return (today - datetime.timedelta(days=int(raw.split()[0]))).strftime("%Y-%m-%d")
        except ValueError:
            pass
    if "week" in raw:
        try:
            return (today - datetime.timedelta(weeks=int(raw.split()[0]))).strftime("%Y-%m-%d")
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%b %d, %Y", "%B %d, %Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def enrich_with_page_data(rows: list[dict]) -> list[dict]:
    """
    Visit each article URL and overwrite date/location with
    accurate values extracted directly from the page.
    """
    for i, row in enumerate(rows):
        url = row.get("Source URL", "")
        if not url or row.get("Source") == "sam_gov":
            continue
        log.info("Fetching page %d/%d: %s", i + 1, len(rows), url)
        meta = fetch_article_meta(url)
        if meta.get("publish_date"):
            row["Date Published"] = meta["publish_date"]
        if meta.get("location"):
            row["Location"] = meta["location"]
        time.sleep(0.5)   # be polite
    return rows


# --- SAM.gov -----------------------------------------------------------------

SAM_API_URL = "https://api.sam.gov/opportunities/v2/search"

def samgov_search(start_date: str, end_date: str) -> list[dict]:
    """
    Search SAM.gov using keyword-based title searches.
    Date format must be MM/DD/YYYY for SAM.gov API.
    """
    # Convert YYYY-MM-DD to MM/DD/YYYY for SAM.gov
    def fmt(d):
        parts = d.split("-")
        return parts[1] + "/" + parts[2] + "/" + parts[0]

    sam_keywords = [
        "hangar",
        "aircraft hangar",
        "hangar fire suppression",
        "hangar fire protection",
        "hangar construction",
        "hangar renovation",
    ]

    results = []
    seen_ids = set()

    for keyword in sam_keywords:
        params = {
            "api_key":    os.environ.get("SAM_API_KEY", "DEMO_KEY"),
            "postedFrom": fmt(start_date),
            "postedTo":   fmt(end_date),
            "title":      keyword,
            "limit":      25,
            "offset":     0,
        }
        try:
            r = requests.get(SAM_API_URL, params=params, timeout=30)
            r.raise_for_status()
            data = r.json()
            opps = data.get("opportunitiesData", [])
            log.info("SAM.gov '%s' -> %d results", keyword, len(opps))
            for opp in opps:
                notice_id = opp.get("noticeId", "")
                if notice_id in seen_ids:
                    continue
                seen_ids.add(notice_id)
                row = parse_sam_result(opp)
                if row:
                    results.append(row)
        except Exception as exc:
            log.warning("SAM.gov error for '%s': %s", keyword, exc)
        time.sleep(0.5)
    return results


def parse_sam_result(opp: dict) -> dict | None:
    title    = opp.get("title", "") or ""
    desc     = opp.get("description", "") or ""
    combined = (title + " " + desc).lower()

    if not any(kw in combined for kw in SAM_KEYWORDS):
        return None

    place        = opp.get("placeOfPerformance", {}) or {}
    state_name   = (place.get("state", {}) or {}).get("name", "")
    country_name = (place.get("country", {}) or {}).get("name", "United States")
    if country_name not in NA_COUNTRIES:
        country_name = "United States"

    # FIX: 'location' was undefined in the previous version, causing a
    # NameError that silently dropped every SAM.gov result.
    if state_name:
        location = country_name + " - " + state_name
    else:
        location = country_name

    posted = opp.get("postedDate", "")[:10] if opp.get("postedDate") else ""
    url    = "https://sam.gov/opp/" + opp.get("noticeId", "") + "/view"

    return {
        "Project Title":  title.strip(),
        "Source URL":     url,
        "Summary":        desc.strip()[:300] if desc else "",
        "Date Published": posted,
        "Location":       location,
        "Source":         "sam_gov",
    }


# --- CanadaBuys (Canada) -----------------------------------------------------
# CanadaBuys publishes daily open data CSV files — no API key required.
# We download the "new tender notices" CSV and filter for hangar keywords.

CANADABUYS_CSV_URL = "https://canadabuys.canada.ca/opendata/pub/newTenderNotice-nouvelAvisAppelOffres.csv"

CANADABUYS_KEYWORDS = [
    "hangar", "aircraft", "airfield", "aviation",
    "fire suppression", "fire protection",
]

def canadabuys_search(start_date: str) -> list[dict]:
    """
    Download CanadaBuys new tender notices CSV and filter for
    hangar/aviation related opportunities published this week.
    """
    results = []
    try:
        r = requests.get(CANADABUYS_CSV_URL, timeout=30)
        r.raise_for_status()

        import csv
        import io as _io
        reader = csv.DictReader(_io.StringIO(r.text))
        for row in reader:
            # Column names vary — try common ones
            title = (
                row.get("title-titre-eng", "") or
                row.get("title_eng", "") or
                row.get("Title", "") or ""
            ).strip()
            pub_date = (
                row.get("publicationDate-datePublication", "") or
                row.get("publication_date", "") or
                row.get("Date", "") or ""
            ).strip()[:10]  # take YYYY-MM-DD part
            notice_id = (
                row.get("referenceNumber-numeroReference", "") or
                row.get("reference_number", "") or
                row.get("ID", "") or ""
            ).strip()
            desc = (
                row.get("description-eng", "") or
                row.get("description", "") or ""
            ).strip()

            if not title:
                continue

            # Filter by keyword
            combined = (title + " " + desc).lower()
            if not any(kw in combined for kw in CANADABUYS_KEYWORDS):
                continue

            # Filter by date — only keep results from this week
            if pub_date and pub_date < start_date:
                continue

            url = (
                "https://canadabuys.canada.ca/en/tender-opportunities/tender-notice/" +
                notice_id if notice_id else "https://canadabuys.canada.ca/en/tender-opportunities"
            )

            results.append({
                "Project Title":  title,
                "Source URL":     url,
                "Summary":        desc[:300] if desc else "",
                "Date Published": pub_date,
                "Region":         "NA",
                "Location":       "Canada",
                "Source":         "canadabuys",
            })

        log.info("CanadaBuys -> %d results", len(results))
    except Exception as exc:
        log.warning("CanadaBuys error: %s", exc)
    return results


# --- AusTender (Australia) ---------------------------------------------------
# AusTender OCDS API — requires a free auth token registered at tenders.gov.au
# Token is passed via the AUSTENDER_TOKEN environment variable / GitHub secret.

AUSTENDER_API_URL = "https://api.tenders.gov.au/atm/v1/releases"

AUSTENDER_KEYWORDS = [
    "hangar", "aircraft", "airfield", "aviation",
    "fire suppression", "fire protection",
]

def austender_search(start_date: str, end_date: str) -> list[dict]:
    """
    Search AusTender OCDS API for hangar-related opportunities.
    Requires AUSTENDER_TOKEN environment variable.
    """
    token = os.environ.get("AUSTENDER_TOKEN", "")
    if not token:
        log.warning("AUSTENDER_TOKEN not set — skipping AusTender search")
        return []

    results = []
    try:
        headers = {
            "Authorization": "Bearer " + token,
            "Accept": "application/json",
        }
        params = {
            "dateFrom": start_date,
            "dateTo":   end_date,
            "limit":    100,
        }
        r = requests.get(AUSTENDER_API_URL, headers=headers, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()

        releases = data.get("releases", [])
        log.info("AusTender raw results: %d", len(releases))

        for release in releases:
            tender = release.get("tender", {}) or {}
            title  = (tender.get("title", "") or "").strip()
            desc   = (tender.get("description", "") or "").strip()

            if not title:
                continue

            combined = (title + " " + desc).lower()
            if not any(kw in combined for kw in AUSTENDER_KEYWORDS):
                continue

            pub_date   = (release.get("date", "") or "")[:10]
            release_id = release.get("ocid", "") or release.get("id", "")
            url = "https://www.tenders.gov.au/Atm/Show/" + release_id if release_id else "https://www.tenders.gov.au"

            results.append({
                "Project Title":  title,
                "Source URL":     url,
                "Summary":        desc[:300] if desc else "",
                "Date Published": pub_date,
                "Location":       "Australia",
                "Source":         "austender",
            })

        log.info("AusTender -> %d relevant results", len(results))
    except Exception as exc:
        log.warning("AusTender error: %s", exc)
    return results


# --- TED Europa (European procurement) ---------------------------------------
# TED (Tenders Electronic Daily) is the EU's official procurement portal.
# The public search API requires no authentication.

TED_API_URL = "https://api.ted.europa.eu/v3/notices/search"

TED_KEYWORDS = [
    "hangar", "aircraft hangar", "fire suppression hangar",
    "hangar fire protection", "hangar construction", "MRO facility",
]

def ted_europa_search(start_date: str, end_date: str) -> list[dict]:
    """
    Search TED Europa for hangar/aviation related procurement notices.
    No API key required for basic public search.
    """
    results = []
    seen_ids = set()

    for keyword in TED_KEYWORDS:
        try:
            payload = {
                "query":  keyword,
                "fields": ["title", "summary", "publicationDate", "noticePublicationId", "noticeType", "buyer-country"],
                "page":   1,
                "limit":  25,
                "filters": {
                    "publicationDate": {
                        "gte": start_date,
                        "lte": end_date,
                    }
                },
                "sort": [{"publicationDate": "desc"}],
            }
            r = requests.post(TED_API_URL, json=payload, timeout=30)
            r.raise_for_status()
            data = r.json()
            notices = data.get("notices", [])
            log.info("TED Europa '%s' -> %d results", keyword, len(notices))

            for notice in notices:
                notice_id = notice.get("noticePublicationId", "")
                if notice_id in seen_ids:
                    continue
                seen_ids.add(notice_id)

                title    = (notice.get("title", {}) or {}).get("eng", "") or ""
                if not title:
                    title = next(iter((notice.get("title", {}) or {}).values()), "")
                summary  = (notice.get("summary", {}) or {}).get("eng", "") or ""
                pub_date = (notice.get("publicationDate", "") or "")[:10]
                url      = "https://ted.europa.eu/en/notice/-/detail/" + notice_id if notice_id else "https://ted.europa.eu"

                if not title:
                    continue

                # Try to detect a more specific country than just "Europe"
                # by scanning title + summary
                detected = detect_location(summary, title=title)
                location = detected if detected else "Europe"

                results.append({
                    "Project Title":  title.strip(),
                    "Source URL":     url,
                    "Summary":        summary.strip()[:300] if summary else "",
                    "Date Published": pub_date,
                    "Location":       location,
                    "Source":         "ted_europa",
                })

        except Exception as exc:
            log.warning("TED Europa error for '%s': %s", keyword, exc)
        time.sleep(0.5)

    log.info("TED Europa total -> %d results", len(results))
    return results


# --- Location detection ------------------------------------------------------
#
# Strategy: score every candidate country based on mentions in the title
# (weighted heavily) and body (weighted lightly). Use word-boundary regex
# to avoid false matches like "India" inside "Indianapolis". Pick the country
# with the highest total score. Country names are stronger signals than city
# names, since cities often appear in unrelated contexts.

US_STATES_FULL = {
    "Alabama","Alaska","Arizona","Arkansas","California","Colorado","Connecticut",
    "Delaware","Florida","Georgia","Hawaii","Idaho","Illinois","Indiana","Iowa",
    "Kansas","Kentucky","Louisiana","Maine","Maryland","Massachusetts","Michigan",
    "Minnesota","Mississippi","Missouri","Montana","Nebraska","Nevada",
    "New Hampshire","New Jersey","New Mexico","New York","North Carolina",
    "North Dakota","Ohio","Oklahoma","Oregon","Pennsylvania","Rhode Island",
    "South Carolina","South Dakota","Tennessee","Texas","Utah","Vermont",
    "Virginia","Washington","West Virginia","Wisconsin","Wyoming",
}

# Two-letter state codes — only matched with strict word boundaries because they
# are short and risk colliding with common words. We keep them but require at
# least one full state name OR a US-specific keyword to confirm "US".
US_STATE_ABBR = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN",
    "IA","KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV",
    "NH","NJ","NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN",
    "TX","UT","VT","VA","WA","WV","WI","WY",
}

CA_PROVINCES_FULL = {
    "Alberta","British Columbia","Manitoba","New Brunswick",
    "Newfoundland and Labrador","Nova Scotia","Ontario",
    "Prince Edward Island","Quebec","Saskatchewan",
    "Northwest Territories","Nunavut","Yukon",
}

MX_STATES = {
    "Aguascalientes","Baja California","Baja California Sur","Campeche",
    "Chiapas","Chihuahua","Coahuila","Colima","Durango","Guanajuato",
    "Guerrero","Hidalgo","Jalisco","Mexico City","Michoacan","Morelos",
    "Nayarit","Nuevo Leon","Oaxaca","Puebla","Queretaro","Quintana Roo",
    "San Luis Potosi","Sinaloa","Sonora","Tabasco","Tamaulipas","Tlaxcala",
    "Veracruz","Yucatan","Zacatecas",
}

# --- Country -> region mapping (canonical) ----------------------------------
# Each country is in EXACTLY one region. No duplicates.

EUROPE_COUNTRIES = {
    "Albania","Andorra","Armenia","Austria","Azerbaijan","Belarus","Belgium",
    "Bosnia and Herzegovina","Bulgaria","Croatia","Cyprus","Czech Republic",
    "Czechia","Denmark","Estonia","Finland","France","Georgia","Germany",
    "Greece","Hungary","Iceland","Ireland","Italy","Kazakhstan","Kosovo",
    "Latvia","Liechtenstein","Lithuania","Luxembourg","Malta","Moldova",
    "Monaco","Montenegro","Netherlands","North Macedonia","Norway","Poland",
    "Portugal","Romania","Russia","San Marino","Serbia","Slovakia","Slovenia",
    "Spain","Sweden","Switzerland","Turkey","Ukraine","United Kingdom",
    "Vatican","England","Scotland","Wales","Northern Ireland",
}

MIDDLE_EAST_COUNTRIES = {
    "Bahrain","Iran","Iraq","Israel","Jordan","Kuwait",
    "Lebanon","Oman","Palestine","Qatar","Saudi Arabia","Syria",
    "United Arab Emirates","Yemen",
}

AFRICA_COUNTRIES = {
    "Algeria","Angola","Benin","Botswana","Burkina Faso","Burundi",
    "Cameroon","Cape Verde","Central African Republic","Chad","Comoros",
    "Congo","Ivory Coast","Djibouti","Egypt","Equatorial Guinea","Eritrea",
    "Eswatini","Ethiopia","Gabon","Gambia","Ghana","Guinea","Guinea-Bissau",
    "Kenya","Lesotho","Liberia","Libya","Madagascar","Malawi","Mali",
    "Mauritania","Mauritius","Morocco","Mozambique","Namibia","Niger",
    "Nigeria","Rwanda","Senegal","Seychelles","Sierra Leone",
    "Somalia","South Africa","South Sudan","Sudan","Tanzania","Togo",
    "Tunisia","Uganda","Zambia","Zimbabwe",
}

# Asia-Pacific — folded into the EMEA tab since the report has only 2 tabs
# and there is no separate APAC tab. Trevor: change this if you want APAC
# to land somewhere else (or to be excluded entirely).
APAC_COUNTRIES = {
    "Afghanistan","Bangladesh","Bhutan","Brunei","Cambodia","China",
    "India","Indonesia","Japan","Laos","Malaysia","Maldives","Mongolia",
    "Myanmar","Nepal","North Korea","Pakistan","Philippines","Singapore",
    "South Korea","Sri Lanka","Taiwan","Thailand","Timor-Leste","Vietnam",
    "Australia","New Zealand","Papua New Guinea","Fiji",
}

# Country aliases / abbreviations — these resolve to canonical country names.
# Keys MUST be unique. Only include unambiguous aliases.
COUNTRY_ALIASES = {
    "USA":               "United States",
    "U.S.":              "United States",
    "U.S.A.":            "United States",
    "America":           "United States",  # weak but common
    "UK":                "United Kingdom",
    "U.K.":              "United Kingdom",
    "Britain":           "United Kingdom",
    "Great Britain":     "United Kingdom",
    "UAE":               "United Arab Emirates",
    "U.A.E.":            "United Arab Emirates",
    "Holland":           "Netherlands",
    "Czech":             "Czechia",
    "South Korean":      "South Korea",
    "North Korean":      "North Korea",
}

# Major cities that map UNAMBIGUOUSLY to a country. Cities that exist in
# multiple countries (London, Birmingham, Paris, Manchester, etc.) are NOT
# included here — relying on the country name to appear elsewhere in the text
# is safer than guessing.
COUNTRY_CITIES = {
    # Middle East — strong, distinctive city signals
    "United Arab Emirates": ["Dubai", "Abu Dhabi", "Sharjah", "Ras Al Khaimah",
                              "Ajman", "Fujairah", "Umm Al Quwain"],
    "Saudi Arabia":         ["Riyadh", "Jeddah", "Dammam", "Mecca", "Medina",
                              "Tabuk", "NEOM"],
    "Qatar":                ["Doha"],
    "Bahrain":              ["Manama"],
    "Kuwait":               ["Kuwait City"],
    "Oman":                 ["Muscat"],
    "Jordan":               ["Amman"],
    "Israel":               ["Tel Aviv", "Jerusalem", "Haifa"],
    "Lebanon":              ["Beirut"],

    # Africa — distinctive city signals
    "Egypt":                ["Cairo", "Alexandria", "Giza"],
    "Nigeria":              ["Lagos", "Abuja", "Port Harcourt", "Ibadan", "Kano"],
    "Kenya":                ["Nairobi", "Mombasa"],
    "South Africa":         ["Johannesburg", "Cape Town", "Pretoria", "Durban"],
    "Morocco":              ["Casablanca", "Rabat", "Marrakech", "Tangier"],
    "Tunisia":              ["Tunis"],
    "Algeria":              ["Algiers"],
    "Ethiopia":             ["Addis Ababa"],
    "Tanzania":             ["Dar es Salaam"],
    "Uganda":               ["Kampala"],
    "Senegal":              ["Dakar"],
    "Ghana":                ["Accra"],

    # Europe — distinctive city signals (skipping ones that collide with US/CA)
    "France":               ["Toulouse", "Marseille", "Lyon", "Nice"],
    "Germany":               ["Berlin", "Munich", "Frankfurt", "Hamburg",
                              "Cologne", "Stuttgart", "Düsseldorf"],
    "Spain":                ["Madrid", "Barcelona", "Valencia", "Seville"],
    "Italy":                ["Rome", "Milan", "Naples", "Turin", "Florence", "Venice"],
    "Netherlands":          ["Amsterdam", "Rotterdam", "Schiphol", "Eindhoven"],
    "Belgium":              ["Brussels", "Antwerp"],
    "Austria":              ["Vienna", "Salzburg"],
    "Poland":               ["Warsaw", "Krakow", "Wroclaw", "Gdansk", "Powidz"],
    "Czechia":              ["Prague", "Brno"],
    "Hungary":              ["Budapest"],
    "Romania":              ["Bucharest"],
    "Greece":               ["Athens", "Thessaloniki"],
    "Portugal":             ["Lisbon", "Porto"],
    "Sweden":               ["Stockholm", "Gothenburg"],
    "Norway":               ["Oslo", "Bergen"],
    "Denmark":              ["Copenhagen"],
    "Finland":              ["Helsinki"],
    "Switzerland":          ["Zurich", "Geneva", "Basel"],
    "Ireland":              ["Dublin"],
    "Cyprus":               ["Nicosia", "Larnaca", "Limassol"],
    "Turkey":               ["Istanbul", "Ankara", "Izmir"],
    "Ukraine":              ["Kyiv", "Kiev", "Lviv"],

    # APAC
    "Australia":            ["Sydney", "Melbourne", "Brisbane", "Perth",
                              "Adelaide", "Canberra"],
    "Japan":                ["Tokyo", "Osaka", "Yokohama", "Nagoya"],
    "Singapore":            ["Changi"],
    "China":                ["Beijing", "Shanghai", "Shenzhen", "Guangzhou"],
    "India":                ["Mumbai", "Bengaluru", "Hyderabad", "Chennai"],
    "Thailand":             ["Bangkok"],
    "Malaysia":             ["Kuala Lumpur"],
    "Indonesia":            ["Jakarta"],
    "Philippines":          ["Manila"],
    "South Korea":          ["Seoul", "Busan", "Incheon"],
    "Vietnam":              ["Hanoi", "Ho Chi Minh"],
    "Taiwan":               ["Taipei"],
}

# Build a master country -> region lookup
COUNTRY_REGION = {}
for c in EUROPE_COUNTRIES:        COUNTRY_REGION[c] = "EMEA"
for c in MIDDLE_EAST_COUNTRIES:   COUNTRY_REGION[c] = "EMEA"
for c in AFRICA_COUNTRIES:        COUNTRY_REGION[c] = "EMEA"
for c in APAC_COUNTRIES:          COUNTRY_REGION[c] = "EMEA"   # see comment above
COUNTRY_REGION["United States"] = "NA"
COUNTRY_REGION["Canada"]         = "NA"
COUNTRY_REGION["Mexico"]         = "NA"

# All canonical country names we'll search for directly
ALL_COUNTRIES = (
    EUROPE_COUNTRIES | MIDDLE_EAST_COUNTRIES | AFRICA_COUNTRIES |
    APAC_COUNTRIES | {"United States", "Canada", "Mexico"}
)


def _word_count(needle: str, haystack_lower: str) -> int:
    """Count whole-word occurrences of needle (case-insensitive)."""
    if not needle:
        return 0
    pattern = r'\b' + re.escape(needle.lower()) + r'\b'
    return len(re.findall(pattern, haystack_lower))


def detect_location(text: str, title: str = "") -> str:
    """
    Detect the most likely location for an article.

    Returns a location string like "United States - Texas", "United Kingdom",
    "Canada - Ontario", "United Arab Emirates", or "" if nothing detected.

    Algorithm: score every candidate country based on weighted mentions in
    title and body, with title hits weighted ~5x heavier. Country-name hits
    are stronger than city-name hits. Returns the highest-scoring country.
    """
    if not text and not title:
        return ""

    title_lower = (title or "").lower()
    body_lower  = (text  or "").lower()

    scores = {}  # canonical country -> float score

    def add_score(country: str, points: float):
        if points <= 0:
            return
        scores[country] = scores.get(country, 0.0) + points

    TITLE_WEIGHT       = 5.0
    BODY_WEIGHT        = 1.0
    CITY_TITLE_WEIGHT  = 4.0
    CITY_BODY_WEIGHT   = 0.7
    ALIAS_TITLE_WEIGHT = 4.0
    ALIAS_BODY_WEIGHT  = 0.8

    # 1. Direct country name matches
    for country in ALL_COUNTRIES:
        t = _word_count(country, title_lower)
        b = _word_count(country, body_lower)
        add_score(country, t * TITLE_WEIGHT + b * BODY_WEIGHT)

    # 2. Country aliases (USA, UK, UAE, Britain, etc.)
    for alias, canonical in COUNTRY_ALIASES.items():
        t = _word_count(alias, title_lower)
        b = _word_count(alias, body_lower)
        add_score(canonical, t * ALIAS_TITLE_WEIGHT + b * ALIAS_BODY_WEIGHT)

    # 3. City name matches (lower weight than country names)
    for country, cities in COUNTRY_CITIES.items():
        for city in cities:
            t = _word_count(city, title_lower)
            b = _word_count(city, body_lower)
            add_score(country, t * CITY_TITLE_WEIGHT + b * CITY_BODY_WEIGHT)

    # 4. UK constituent countries map to "United Kingdom" for region purposes
    for uk_part in ("England", "Scotland", "Wales", "Northern Ireland"):
        t = _word_count(uk_part, title_lower)
        b = _word_count(uk_part, body_lower)
        if t or b:
            add_score("United Kingdom", t * TITLE_WEIGHT + b * BODY_WEIGHT)

    # 5. US state matching → contributes to "United States" with state name
    us_state_score = 0.0
    detected_state = None
    detected_state_score = 0.0
    for st in US_STATES_FULL:
        t = _word_count(st, title_lower)
        b = _word_count(st, body_lower)
        s = t * TITLE_WEIGHT + b * BODY_WEIGHT
        if s > 0:
            us_state_score += s
            # Track the highest-scoring state for the output
            if s > detected_state_score:
                detected_state_score = s
                detected_state = st
    if us_state_score > 0:
        add_score("United States", us_state_score)

    # 6. Canadian province matching → contributes to "Canada"
    ca_score = 0.0
    detected_province = None
    detected_province_score = 0.0
    for prov in CA_PROVINCES_FULL:
        t = _word_count(prov, title_lower)
        b = _word_count(prov, body_lower)
        s = t * TITLE_WEIGHT + b * BODY_WEIGHT
        if s > 0:
            ca_score += s
            if s > detected_province_score:
                detected_province_score = s
                detected_province = prov
    if ca_score > 0:
        add_score("Canada", ca_score)

    # 7. Mexican state matching → contributes to "Mexico"
    mx_score = 0.0
    for st in MX_STATES:
        t = _word_count(st, title_lower)
        b = _word_count(st, body_lower)
        mx_score += t * TITLE_WEIGHT + b * BODY_WEIGHT
    if mx_score > 0:
        add_score("Mexico", mx_score)

    # No signals at all
    if not scores:
        return ""

    # Pick the winner
    best_country = max(scores, key=scores.get)

    # Format output with state/province if applicable
    if best_country == "United States" and detected_state:
        return "United States - " + detected_state
    if best_country == "Canada" and detected_province:
        return "Canada - " + detected_province
    return best_country


# --- Region classification ---------------------------------------------------

def classify_region(location: str) -> str:
    """
    Returns "NA" or "EMEA" for tab assignment.

    Empty/unknown locations default to NA (since most NA-source feeds —
    SAM.gov, CanadaBuys, US-biased Google search — produce them).
    """
    if not location:
        return "NA"

    # Strip "Country - State/Province" formatting to get just the country
    base = location.split(" - ", 1)[0].strip()

    # Direct lookup in the region map
    region = COUNTRY_REGION.get(base)
    if region:
        return region

    # Fallback partial-match lookup (handles edge cases like "Europe", "EMEA")
    base_lower = base.lower()
    if base_lower in ("europe", "european union", "eu", "emea", "middle east",
                      "africa", "asia", "asia pacific", "apac"):
        return "EMEA"
    if base_lower in ("north america", "americas"):
        return "NA"

    # Unknown — default to NA so we don't dump random international stuff
    # into EMEA either; better to err on the side that maintainers will notice
    return "NA"


def split_by_region(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Split a combined list of rows into NA and EMEA based on detected location.
    """
    na_rows   = []
    emea_rows = []

    for row in rows:
        location = row.get("Location", "") or ""
        region   = classify_region(location)
        if region == "EMEA":
            emea_rows.append(row)
        else:
            na_rows.append(row)

    return na_rows, emea_rows

# --- Deduplication -----------------------------------------------------------

def deduplicate(rows: list[dict]) -> list[dict]:
    seen_urls, seen_titles, unique = set(), set(), []
    for row in rows:
        url   = (row.get("Source URL") or "").strip().lower()
        title = (row.get("Project Title") or "").strip().lower()
        if (url and url in seen_urls) or (title and title in seen_titles):
            continue
        seen_urls.add(url)
        seen_titles.add(title)
        unique.append(row)
    return unique


# --- Excel report ------------------------------------------------------------

HEADERS = [
    "Project Title", "Source URL", "Summary",
    "Date Published", "Location",
]

COL_WIDTHS = {
    "Project Title":  40, "Source URL": 50, "Summary": 60,
    "Date Published": 15, "Location": 25,
}

HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="DCE6F1")
THIN         = Side(style="thin", color="AAAAAA")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, rows: list[dict]):
    """Write a set of rows into a given worksheet with full formatting."""
    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(HEADERS, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER

    ws.row_dimensions[1].height = 20

    for row_idx, row in enumerate(rows, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, header in enumerate(HEADERS, start=1):
            value          = row.get(header, "")
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font      = Font(name="Calibri", size=10)
            if header == "Source URL" and value:
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    ws.auto_filter.ref = ws.dimensions


def build_workbook(na_rows: list[dict], emea_rows: list[dict], week: str) -> openpyxl.Workbook:
    """Build a workbook with two tabs: North America and EMEA."""
    wb = openpyxl.Workbook()

    # Tab 1: North America
    ws_na       = wb.active
    ws_na.title = "North America - " + week
    write_sheet(ws_na, na_rows)

    # Tab 2: EMEA
    ws_emea       = wb.create_sheet()
    ws_emea.title = "EMEA - " + week
    write_sheet(ws_emea, emea_rows)

    return wb


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# --- Email -------------------------------------------------------------------

EMAIL_BODY = """\
Hi Safespill Team,
{notice_block}
Please find attached this week's Safespill Hangar Intelligence Report.

The report covers new hangar projects, retrofit work, construction contracts,
and fire protection opportunities globally, discovered between
{start_date} and {end_date}.

The Excel file contains two tabs:
  - North America: Google Search, SAM.gov, CanadaBuys
  - EMEA: Google Search, TED Europa, AusTender (Australia/APAC folded into EMEA)

Total results this week: {count}

Safespill Automated Intelligence Report
"""


def build_quota_notice(ending_quota: dict, error_count: int) -> str:
    """
    Build the optional notice paragraph that appears in the email body when
    SerpAPI quota is low or rate-limit errors occurred during the run.
    Returns an empty string in normal weeks (no notice is shown).
    """
    # Rate-limit / quota errors occurred during this run — most severe case
    if error_count > 0:
        return (
            "WARNING: Coverage may have been incomplete this week. The scraper "
            "encountered SerpAPI rate or quota limits {n} time(s) during this "
            "run, meaning some search queries returned no results. Upgrading "
            "from the free SerpAPI plan ($25/month for 4x capacity) would "
            "resolve this."
        ).format(n=error_count)

    if not ending_quota:
        return ""

    try:
        left    = int(ending_quota.get("total_searches_left", 0))
        monthly = int(ending_quota.get("searches_per_month", 0))
        if monthly <= 0:
            return ""
        pct_left = 100.0 * left / monthly
    except (ValueError, TypeError):
        return ""

    if pct_left < 10:
        return (
            "WARNING: SerpAPI quota is critically low ({left} of {monthly} "
            "searches remaining for the month). Next week's report may be "
            "incomplete unless the API plan is upgraded."
        ).format(left=left, monthly=monthly)

    if pct_left < 25:
        pct_used = 100 - pct_left
        return (
            "Note: SerpAPI usage is at {pct:.0f}% for the month ({left} "
            "searches remaining). Coverage may be reduced in the final run "
            "of this billing cycle."
        ).format(pct=pct_used, left=left)

    # Normal week — no notice
    return ""


def send_email(xlsx_bytes: bytes, filename: str, start_date: str, end_date: str,
               count: int, quota_notice: str = ""):
    msg            = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = RECIPIENT
    msg["Subject"] = "Safespill Hangar Intelligence Report - Week of " + week_label()

    # The notice block is a blank line in normal weeks, or a notice paragraph
    # surrounded by blank lines when there's something worth flagging.
    notice_block = ""
    if quota_notice:
        notice_block = "\n" + quota_notice + "\n"

    msg.attach(MIMEText(EMAIL_BODY.format(
        notice_block=notice_block,
        start_date=start_date, end_date=end_date, count=count), "plain"))

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(xlsx_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", 'attachment; filename="' + filename + '"')
    msg.attach(part)

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        recipient_list = [r.strip() for r in RECIPIENT.split(",")]
        server.sendmail(SMTP_USER, recipient_list, msg.as_string())

    log.info("Email sent to %s", RECIPIENT)


# --- Main --------------------------------------------------------------------

def main():
    global _SERPAPI_ERROR_COUNT
    _SERPAPI_ERROR_COUNT = 0

    start_date, end_date = date_range()
    week = week_label()
    log.info("Running report for %s -> %s", start_date, end_date)

    # Check starting SerpAPI quota (free call — does NOT count against budget)
    starting_quota = check_serpapi_quota()
    log_quota_status("Start", starting_quota)

    # ---- Collect all results into one pool ----
    all_rows = []

    # 1. Google Search — general queries (region auto-detected from results)
    for query in GENERAL_QUERIES:
        log.info("[Search] '%s'", query)
        items = serpapi_news_search(query)
        for item in items:
            row = parse_serpapi_result(item)
            if row:
                all_rows.append(row)
        time.sleep(1)

    # 2. SAM.gov (uses longer 30-day lookback to catch infrequently posted contracts)
    sam_end   = datetime.date.today()
    sam_start = sam_end - datetime.timedelta(days=SAM_LOOKBACK_DAYS)
    sam_start_str = sam_start.strftime("%Y-%m-%d")
    sam_end_str   = sam_end.strftime("%Y-%m-%d")
    log.info("[NA] Searching SAM.gov (past %d days) ...", SAM_LOOKBACK_DAYS)
    all_rows.extend(samgov_search(sam_start_str, sam_end_str))

    # 3. CanadaBuys
    log.info("[NA] Searching CanadaBuys ...")
    all_rows.extend(canadabuys_search(start_date))

    # 4. AusTender (Australia — folded into EMEA tab)
    log.info("[APAC] Searching AusTender ...")
    all_rows.extend(austender_search(start_date, end_date))

    # 5. Google Search — region-targeted queries (with country/city names)
    for query in REGIONAL_QUERIES:
        log.info("[Regional Search] '%s'", query)
        items = serpapi_news_search(query)
        for item in items:
            row = parse_serpapi_result(item)
            if row:
                all_rows.append(row)
        time.sleep(1)

    # 6. TED Europa
    log.info("[EMEA] Searching TED Europa ...")
    all_rows.extend(ted_europa_search(start_date, end_date))

    # 7. Deduplicate entire pool
    all_rows = deduplicate(all_rows)
    log.info("Total unique before enrichment: %d", len(all_rows))

    # 8. Enrich all articles with accurate dates and locations
    log.info("Fetching article pages for accurate dates and locations ...")
    all_rows = enrich_with_page_data(all_rows)

    # 9. Split into NA and EMEA based on detected location
    na_rows, emea_rows = split_by_region(all_rows)
    log.info("After location split — NA: %d | EMEA: %d", len(na_rows), len(emea_rows))

    # 10. Sort by date
    na_rows.sort(key=lambda r: r.get("Date Published", ""), reverse=True)
    emea_rows.sort(key=lambda r: r.get("Date Published", ""), reverse=True)

    total = len(na_rows) + len(emea_rows)
    log.info("Total unique results: %d (NA: %d, EMEA: %d)", total, len(na_rows), len(emea_rows))

    # 11. Build Excel with two tabs
    wb       = build_workbook(na_rows, emea_rows, week)
    xlsx     = workbook_to_bytes(wb)
    filename = "Safespill_Hangar_Report_" + week + ".xlsx"

    # Check ending SerpAPI quota BEFORE sending email so we can include
    # warnings in the email body if quota is low or errors occurred.
    ending_quota = check_serpapi_quota()
    log_quota_status("End", ending_quota)

    quota_notice = build_quota_notice(ending_quota, _SERPAPI_ERROR_COUNT)

    # 12. Send email (with optional quota warning notice)
    send_email(xlsx, filename, start_date, end_date, total, quota_notice)

    if starting_quota and ending_quota:
        try:
            used_this_run = (
                int(starting_quota.get("total_searches_left", 0)) -
                int(ending_quota.get("total_searches_left", 0))
            )
            left_after = int(ending_quota.get("total_searches_left", 0))
            monthly    = int(ending_quota.get("searches_per_month", 0))
            log.info(
                "SerpAPI usage this run: %d searches | %d remaining of %d/month "
                "| %d rate-limit/quota errors during run",
                used_this_run, left_after, monthly, _SERPAPI_ERROR_COUNT,
            )
            # Warn in logs if quota is running low — gives Trevor a heads-up
            # to either reduce queries or ask the CEO about a paid plan.
            if monthly > 0:
                pct_left = 100.0 * left_after / monthly
                if pct_left < 10:
                    log.warning(
                        "SerpAPI quota CRITICAL — only %d searches (%.0f%%) left "
                        "for the month. Next run may fail. Time to consider upgrading.",
                        left_after, pct_left,
                    )
                elif pct_left < 25:
                    log.warning(
                        "SerpAPI quota LOW — %d searches (%.0f%%) left for the month.",
                        left_after, pct_left,
                    )
        except (ValueError, TypeError):
            pass

    log.info("Done.")


if __name__ == "__main__":
    main()

